from __future__ import annotations

import asyncio
import logging
import re
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any

import httpx
from openpyxl import load_workbook

import app.services.postgres_service as db
from app.core.config import settings
from app.services.elevenlabs_service import (
    get_conversation,
    get_conversation_by_call_sid,
    initiate_outbound_call,
)

logger = logging.getLogger(__name__)

_VALID_BLOOD_TYPES = {"A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"}
_NOMINATIM_SEARCH_URL = "https://nominatim.openstreetmap.org/search"
_LAT_LON_PATTERN = re.compile(
    r"^\s*([+-]?\d{1,2}(?:\.\d+)?)\s*[,\s]\s*([+-]?\d{1,3}(?:\.\d+)?)\s*$"
)


def _safe_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _tokenize_location(value: str) -> set[str]:
    return {
        token
        for token in re.split(r"[^a-z0-9]+", value.lower())
        if len(token) >= 3
    }


def _score_geocode_result(result: dict[str, Any], query: str) -> float:
    query_tokens = _tokenize_location(query)
    display_name = str(result.get("display_name") or "")
    display_tokens = _tokenize_location(display_name)

    overlap = 0.0
    if query_tokens:
        overlap = len(query_tokens & display_tokens) / len(query_tokens)

    importance = _safe_float(result.get("importance")) or 0.0
    place_rank = (_safe_float(result.get("place_rank")) or 0.0) / 30.0
    country_code = str(result.get("address", {}).get("country_code") or "").lower()
    india_bonus = 0.15 if country_code == "in" else 0.0

    # Weight text overlap highest so we avoid wrong cities with high global importance.
    return (overlap * 0.65) + (importance * 0.2) + (place_rank * 0.15) + india_bonus


def _candidate_location_queries(raw_query: str) -> list[str]:
    query = re.sub(r"\s+", " ", str(raw_query or "")).strip(" ,")
    if not query:
        return []

    candidates: list[str] = [query]
    lower_query = query.lower()

    if "india" not in lower_query:
        candidates.append(f"{query}, India")

    # Remove obvious floor/unit words that can hurt geocoder precision.
    stripped = re.sub(
        r"\b(\d+(st|nd|rd|th)?\s+floor|floor\s+\d+|building|bldg|unit\s+\d+)\b",
        "",
        query,
        flags=re.IGNORECASE,
    )
    stripped = re.sub(r"\s+", " ", stripped).strip(" ,")
    if stripped and stripped not in candidates:
        candidates.append(stripped)
        if "india" not in stripped.lower():
            candidates.append(f"{stripped}, India")

    # Preserve order, remove empties/dupes.
    deduped: list[str] = []
    seen: set[str] = set()
    for item in candidates:
        if item and item not in seen:
            seen.add(item)
            deduped.append(item)
    return deduped


@lru_cache(maxsize=1024)
def _geocode_plain_location(location_text: str) -> tuple[float | None, float | None]:
    query = str(location_text or "").strip()
    if not query:
        return None, None

    try:
        with httpx.Client(timeout=3.5) as client:
            headers = {
                # Nominatim requires a non-default identifying user-agent.
                "User-Agent": "CareSync-AI/1.0 (blood-campaign-geocoder)",
                "Accept-Language": "en",
            }
            best_match: tuple[float, float, float] | None = None

            for candidate_query in _candidate_location_queries(query):
                for pass_params in (
                    # Prefer India results for this project.
                    {
                        "q": candidate_query,
                        "format": "jsonv2",
                        "addressdetails": 1,
                        "countrycodes": "in",
                        "limit": 5,
                    },
                    # Fallback to global search if India-only returns none.
                    {
                        "q": candidate_query,
                        "format": "jsonv2",
                        "addressdetails": 1,
                        "limit": 5,
                    },
                ):
                    response = client.get(
                        _NOMINATIM_SEARCH_URL,
                        params=pass_params,
                        headers=headers,
                    )
                    if response.status_code >= 400:
                        continue

                    payload = response.json()
                    if not isinstance(payload, list) or not payload:
                        continue

                    for row in payload:
                        if not isinstance(row, dict):
                            continue
                        lat = _safe_float(row.get("lat"))
                        lon = _safe_float(row.get("lon"))
                        if lat is None or lon is None:
                            continue
                        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
                            continue

                        score = _score_geocode_result(row, candidate_query)
                        if not best_match or score > best_match[2]:
                            best_match = (lat, lon, score)

                    # Good enough threshold: keep latency low once we have a strong match.
                    if best_match and best_match[2] >= 0.75:
                        return best_match[0], best_match[1]

            if best_match:
                return best_match[0], best_match[1]

            return None, None
    except Exception:
        return None, None


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _coerce_phone_text(raw_phone: Any) -> str:
    if raw_phone is None:
        return ""

    if isinstance(raw_phone, bool):
        return ""

    if isinstance(raw_phone, int):
        return str(raw_phone)

    if isinstance(raw_phone, float):
        if not raw_phone.is_integer():
            raise ValueError("phone number must not contain decimals")
        return str(int(raw_phone))

    text = str(raw_phone).strip()
    if not text:
        return ""

    numeric_candidate = text.replace(",", "")
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", numeric_candidate):
        try:
            decimal_value = Decimal(numeric_candidate)
            if decimal_value == decimal_value.to_integral_value():
                return format(decimal_value.quantize(Decimal("1")), "f")
        except InvalidOperation:
            pass

    return text


def _normalize_phone(raw_phone: Any) -> str:
    phone_text = _coerce_phone_text(raw_phone)
    digits = "".join(ch for ch in phone_text if ch.isdigit())
    if len(digits) < 10:
        raise ValueError("phone number must contain at least 10 digits")

    # Common local format: 0XXXXXXXXXX
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]

    local_part = digits[-10:]
    return f"+91{local_part}"


def _normalize_gender(raw_gender: str) -> str:
    value = str(raw_gender or "").strip().lower()
    if value in {"male", "m"}:
        return "male"
    if value in {"female", "f"}:
        return "female"
    raise ValueError("gender must be male/female (or m/f)")


def _normalize_blood_type(raw_blood_type: str) -> str:
    value = str(raw_blood_type or "").strip().upper().replace(" ", "")
    if value not in _VALID_BLOOD_TYPES:
        raise ValueError(
            "blood_type must be one of A+, A-, B+, B-, AB+, AB-, O+, O-"
        )
    return value


def _parse_excel_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value or "").strip()
    if not text:
        raise ValueError("last_donated_date is required")

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    raise ValueError("last_donated_date must be a valid date")


def _parse_location_coordinates(location: str) -> tuple[float | None, float | None]:
    text = str(location or "").strip()
    if not text:
        return None, None

    match = _LAT_LON_PATTERN.match(text)
    if not match:
        return _geocode_plain_location(text)

    lat = _safe_float(match.group(1))
    lon = _safe_float(match.group(2))
    if lat is None or lon is None:
        return _geocode_plain_location(text)

    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
        return _geocode_plain_location(text)

    return lat, lon


def _get_sheet_by_name(workbook, expected_name: str):
    for sheet in workbook.worksheets:
        if sheet.title.strip().lower() == expected_name:
            return sheet
    return None


def _header_map(row_values: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, value in enumerate(row_values):
        key = str(value or "").strip().lower().replace(" ", "_")
        if key:
            mapping[key] = idx
    return mapping


def _row_value(row: tuple[Any, ...], headers: dict[str, int], key: str) -> Any:
    idx = headers.get(key)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def _row_text(row: tuple[Any, ...], headers: dict[str, int], key: str) -> str:
    return str(_row_value(row, headers, key) or "").strip()


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(str(cell or "").strip() == "" for cell in row)


def _parse_donor_row(row: tuple[Any, ...], headers: dict[str, int]) -> dict[str, Any]:
    name = _row_text(row, headers, "name")
    gender = _normalize_gender(_row_value(row, headers, "gender"))
    phone = _normalize_phone(_row_value(row, headers, "phone_number"))
    location = _row_text(row, headers, "location")
    donated_date = _parse_excel_date(_row_value(row, headers, "last_donated_date"))
    blood_type = _normalize_blood_type(_row_value(row, headers, "blood_type"))

    if not name:
        raise ValueError("name is required")
    if not location:
        raise ValueError("location is required")

    lat, lon = _parse_location_coordinates(location)

    return {
        "name": name,
        "gender": gender,
        "phone_number": phone,
        "location": location,
        "latitude": lat,
        "longitude": lon,
        "last_donated_date": donated_date,
        "blood_type": blood_type,
    }


def _parse_ngo_row(
    row: tuple[Any, ...],
    headers: dict[str, int],
    *,
    fallback_name_key: str | None = None,
) -> dict[str, Any]:
    ngo_name = _row_text(row, headers, "ngo_name")
    if not ngo_name and fallback_name_key:
        ngo_name = _row_text(row, headers, fallback_name_key)

    phone = _normalize_phone(_row_value(row, headers, "phone_number"))
    location = _row_text(row, headers, "location")

    if not ngo_name:
        raise ValueError("ngo_name is required")
    if not location:
        raise ValueError("location is required")

    lat, lon = _parse_location_coordinates(location)

    return {
        "ngo_name": ngo_name,
        "phone_number": phone,
        "location": location,
        "latitude": lat,
        "longitude": lon,
    }


def _append_rejected(
    rejected_rows: list[dict[str, Any]],
    *,
    sheet_name: str,
    row_index: int,
    exc: Exception,
) -> None:
    rejected_rows.append(
        {
            "sheet": sheet_name,
            "row": row_index,
            "error": str(exc),
        }
    )


def parse_blood_excel(contents: bytes) -> dict[str, Any]:
    workbook = load_workbook(filename=BytesIO(contents), data_only=True)

    donor_sheet = _get_sheet_by_name(workbook, "donors")
    ngo_sheet = _get_sheet_by_name(workbook, "ngos")

    donor_required = {
        "name",
        "gender",
        "phone_number",
        "location",
        "last_donated_date",
        "blood_type",
    }
    ngo_required = {"ngo_name", "phone_number", "location"}

    donors: list[dict[str, Any]] = []
    ngos: list[dict[str, Any]] = []
    rejected_rows: list[dict[str, Any]] = []

    # Format A: Existing 2-sheet contract (Donors + NGOs)
    if donor_sheet is not None and ngo_sheet is not None:
        donor_rows = list(donor_sheet.iter_rows(values_only=True))
        ngo_rows = list(ngo_sheet.iter_rows(values_only=True))

        if not donor_rows:
            raise ValueError("Donors sheet is empty")
        if not ngo_rows:
            raise ValueError("NGOs sheet is empty")

        donor_headers = _header_map(list(donor_rows[0]))
        ngo_headers = _header_map(list(ngo_rows[0]))

        missing_donor = sorted(donor_required - set(donor_headers.keys()))
        missing_ngo = sorted(ngo_required - set(ngo_headers.keys()))
        if missing_donor:
            raise ValueError(f"Donors sheet missing columns: {', '.join(missing_donor)}")
        if missing_ngo:
            raise ValueError(f"NGOs sheet missing columns: {', '.join(missing_ngo)}")

        for row_index, row in enumerate(donor_rows[1:], start=2):
            if _is_blank_row(row):
                continue
            try:
                donors.append(_parse_donor_row(row, donor_headers))
            except Exception as exc:
                _append_rejected(
                    rejected_rows,
                    sheet_name="Donors",
                    row_index=row_index,
                    exc=exc,
                )

        for row_index, row in enumerate(ngo_rows[1:], start=2):
            if _is_blank_row(row):
                continue
            try:
                ngos.append(_parse_ngo_row(row, ngo_headers))
            except Exception as exc:
                _append_rejected(
                    rejected_rows,
                    sheet_name="NGOs",
                    row_index=row_index,
                    exc=exc,
                )

        return {
            "donors": donors,
            "ngos": ngos,
            "rejected_rows": rejected_rows,
        }

    # Format B: Single-sheet mixed rows.
    # Donor rows: require donor columns.
    # NGO rows: donor-only fields may be empty; ngo_name OR name is used as NGO name.
    if len(workbook.worksheets) == 1:
        single_sheet = workbook.worksheets[0]
        single_rows = list(single_sheet.iter_rows(values_only=True))
        if not single_rows:
            raise ValueError("Single sheet is empty")

        headers = _header_map(list(single_rows[0]))
        missing_single = sorted(donor_required - set(headers.keys()))
        if missing_single:
            raise ValueError(
                "Single sheet missing donor columns: " + ", ".join(missing_single)
            )
        if "phone_number" not in headers or "location" not in headers:
            raise ValueError("Single sheet requires phone_number and location columns")

        for row_index, row in enumerate(single_rows[1:], start=2):
            if _is_blank_row(row):
                continue

            row_type = _row_text(row, headers, "row_type").lower()
            ngo_name = _row_text(row, headers, "ngo_name")
            name_value = _row_text(row, headers, "name")
            has_donor_only_values = any(
                _row_text(row, headers, key)
                for key in ("gender", "last_donated_date", "blood_type")
            )

            treat_as_ngo = (
                row_type in {"ngo", "ngo_row"}
                or bool(ngo_name)
                or (bool(name_value) and not has_donor_only_values)
            )

            try:
                if treat_as_ngo:
                    ngos.append(
                        _parse_ngo_row(
                            row,
                            headers,
                            fallback_name_key="name",
                        )
                    )
                else:
                    donors.append(_parse_donor_row(row, headers))
            except Exception as exc:
                _append_rejected(
                    rejected_rows,
                    sheet_name=single_sheet.title,
                    row_index=row_index,
                    exc=exc,
                )

        return {
            "donors": donors,
            "ngos": ngos,
            "rejected_rows": rejected_rows,
        }

    raise ValueError(
        "Excel must be either: (a) two sheets named Donors and NGOs, or "
        "(b) a single sheet with mixed rows"
    )


def ingest_blood_excel(doctor_id: str, contents: bytes) -> dict[str, Any]:
    parsed = parse_blood_excel(contents)

    donor_result = db.upsert_blood_donors(doctor_id, parsed["donors"])
    ngo_result = db.upsert_blood_ngos(doctor_id, parsed["ngos"])

    return {
        "donors": {
            "accepted": len(parsed["donors"]),
            "created": donor_result["created"],
            "updated": donor_result["updated"],
        },
        "ngos": {
            "accepted": len(parsed["ngos"]),
            "created": ngo_result["created"],
            "updated": ngo_result["updated"],
        },
        "rejected_rows": parsed["rejected_rows"],
    }


def _dcr_value(results: dict[str, Any], key: str) -> str:
    raw = results.get(key)
    if isinstance(raw, dict):
        value = raw.get("value")
    else:
        value = raw
    return str(value or "").strip()


def _is_positive_outcome(conversation: dict[str, Any]) -> bool:
    analysis = conversation.get("analysis") or {}
    dcr = analysis.get("data_collection_results") or {}

    donor_consent = _dcr_value(dcr, "donor_consent").lower()
    call_outcome = _dcr_value(dcr, "call_outcome").lower()

    transcript = conversation.get("transcript", "")
    transcript_text = transcript if isinstance(transcript, str) else str(transcript)
    transcript_text = transcript_text.lower()

    if donor_consent in {"yes", "true", "1", "accepted"}:
        return True
    if any(keyword in call_outcome for keyword in ("accepted", "yes", "confirmed")):
        return True
    if "yes" in transcript_text and "donat" in transcript_text:
        return True

    return False


async def _process_donor_attempt(
    *,
    campaign: dict[str, Any],
    donor: dict[str, Any],
    attempt: dict[str, Any],
    worker: dict[str, Any],
) -> dict[str, Any]:
    attempt_id = str(attempt["id"])

    try:
        db.update_campaign_attempt(
            attempt_id,
            {
                "status": "active",
                "started_at": _utc_now_iso(),
                "updated_at": _utc_now_iso(),
            },
        )

        call_reason = (
            f"Urgent blood requirement for {campaign.get('recipient_name', 'a patient')}"
        )
        if campaign.get("reason"):
            call_reason = f"{call_reason}. Reason: {campaign['reason']}"

        call_result = await initiate_outbound_call(
            patient_phone=donor["phone_number"],
            patient_name=donor["name"],
            doctor_name="CareSync Blood Support",
            doctor_specialty="Emergency Blood Coordination",
            lab_result_summary=(
                f"Required blood type: {campaign['blood_type']}"
            ),
            facility_name="Nearest NGO Blood Support Network",
            facility_address=campaign.get("patient_location") or "",
            facility_phone_number="",
            call_reason=call_reason,
            available_slots="Please confirm if you can donate as soon as possible.",
            report_title="Blood Donation Request",
            report_date=datetime.now(timezone.utc).date().isoformat(),
            extra_context={
                "campaign_id": str(campaign["id"]),
                "attempt_id": attempt_id,
                "donor_id": str(donor["id"]),
            },
            agent_id_override=worker.get("agent_id"),
            phone_number_id_override=worker.get("phone_number_id"),
        )

        conversation_id = call_result.get("conversation_id") or ""
        call_sid = call_result.get("callSid") or call_result.get("call_sid") or ""

        db.update_campaign_attempt(
            attempt_id,
            {
                "conversation_id": conversation_id or None,
                "call_sid": call_sid or None,
                "message": "Call initiated",
                "updated_at": _utc_now_iso(),
            },
        )

        resolved_conversation_id = conversation_id
        max_polls = 40
        poll_interval_sec = 10

        for _ in range(max_polls):
            if not resolved_conversation_id and call_sid:
                resolved_conversation_id = await get_conversation_by_call_sid(call_sid)
                if resolved_conversation_id:
                    db.update_campaign_attempt(
                        attempt_id,
                        {
                            "conversation_id": resolved_conversation_id,
                            "updated_at": _utc_now_iso(),
                        },
                    )

            if not resolved_conversation_id:
                await asyncio.sleep(poll_interval_sec)
                continue

            conversation = await get_conversation(resolved_conversation_id)
            status = str(conversation.get("status") or "").lower()
            if status in {"in_progress", "processing", "initiated", "queued"}:
                await asyncio.sleep(poll_interval_sec)
                continue

            accepted = _is_positive_outcome(conversation)
            new_status = "accepted" if accepted else "completed"
            db.update_campaign_attempt(
                attempt_id,
                {
                    "status": new_status,
                    "outcome": "accepted" if accepted else (status or "completed"),
                    "message": "Donor accepted" if accepted else "Call completed",
                    "completed_at": _utc_now_iso(),
                    "updated_at": _utc_now_iso(),
                },
            )
            return {
                "accepted": accepted,
                "donor_id": str(donor["id"]),
                "attempt_id": attempt_id,
            }

        db.update_campaign_attempt(
            attempt_id,
            {
                "status": "failed",
                "outcome": "timeout",
                "message": "Timed out waiting for call completion",
                "completed_at": _utc_now_iso(),
                "updated_at": _utc_now_iso(),
            },
        )
        return {"accepted": False, "donor_id": str(donor["id"]), "attempt_id": attempt_id}

    except Exception as exc:
        logger.exception("Blood campaign donor attempt failed")
        db.update_campaign_attempt(
            attempt_id,
            {
                "status": "failed",
                "outcome": "error",
                "message": str(exc),
                "completed_at": _utc_now_iso(),
                "updated_at": _utc_now_iso(),
            },
        )
        return {"accepted": False, "donor_id": str(donor["id"]), "attempt_id": attempt_id}


async def run_blood_campaign_dispatch(campaign_id: str) -> None:
    campaign = db.get_blood_campaign(campaign_id)
    if not campaign:
        return

    db.update_blood_campaign(
        campaign_id,
        {
            "status": "running",
            "started_at": _utc_now_iso(),
            "updated_at": _utc_now_iso(),
        },
    )

    batch_size = max(1, int(campaign.get("batch_size") or 3))

    while True:
        latest_campaign = db.get_blood_campaign(campaign_id)
        if not latest_campaign:
            return

        if latest_campaign.get("status") not in {"running", "draft"}:
            return

        donors = db.list_campaign_eligible_donors(campaign_id, limit=batch_size, offset=0)
        if not donors:
            db.update_blood_campaign(
                campaign_id,
                {
                    "status": "completed",
                    "completed_at": _utc_now_iso(),
                    "updated_at": _utc_now_iso(),
                },
            )
            return

        worker_pools = db.list_call_worker_pools(str(latest_campaign["doctor_id"]))
        if not worker_pools:
            worker_pools = [
                {
                    "id": None,
                    "worker_name": f"default-worker-{i + 1}",
                    "agent_id": settings.elevenlabs_agent_id,
                    "phone_number_id": settings.elevenlabs_phone_number_id,
                }
                for i in range(batch_size)
                if settings.elevenlabs_agent_id and settings.elevenlabs_phone_number_id
            ]

        if not worker_pools:
            db.update_blood_campaign(
                campaign_id,
                {
                    "status": "failed",
                    "completed_at": _utc_now_iso(),
                    "updated_at": _utc_now_iso(),
                },
            )
            return

        tasks = []
        for idx, donor in enumerate(donors):
            worker = worker_pools[idx % len(worker_pools)]
            attempt = db.create_campaign_attempt(
                {
                    "campaign_id": campaign_id,
                    "donor_id": donor["id"],
                    "worker_pool_id": worker.get("id"),
                    "status": "queued",
                    "eligibility_reason": "blood type + cooldown eligible",
                    "message": "Queued for call",
                }
            )
            tasks.append(
                _process_donor_attempt(
                    campaign=latest_campaign,
                    donor=donor,
                    attempt=attempt,
                    worker=worker,
                )
            )

        batch_results = await asyncio.gather(*tasks, return_exceptions=False)
        accepted = [result for result in batch_results if result.get("accepted")]

        if accepted:
            accepted_donor_id = accepted[0]["donor_id"]
            db.update_blood_campaign(
                campaign_id,
                {
                    "status": "stopped",
                    "accepted_donor_id": accepted_donor_id,
                    "completed_at": _utc_now_iso(),
                    "updated_at": _utc_now_iso(),
                },
            )
            return


async def start_blood_campaign(
    *,
    doctor_id: str,
    blood_type: str,
    recipient_name: str,
    reason: str | None,
    patient_location: str | None,
    batch_size: int,
) -> dict[str, Any]:
    patient_latitude, patient_longitude = _parse_location_coordinates(patient_location or "")

    campaign = db.create_blood_campaign(
        {
            "doctor_id": doctor_id,
            "blood_type": _normalize_blood_type(blood_type),
            "recipient_name": recipient_name.strip(),
            "reason": (reason or "").strip() or None,
            "patient_location": (patient_location or "").strip() or None,
            "patient_latitude": patient_latitude,
            "patient_longitude": patient_longitude,
            "batch_size": max(1, int(batch_size or 3)),
            "status": "draft",
        }
    )

    asyncio.create_task(run_blood_campaign_dispatch(str(campaign["id"])))
    return campaign


def get_blood_campaign_snapshot(campaign_id: str) -> dict[str, Any] | None:
    campaign = db.get_blood_campaign(campaign_id)
    if not campaign:
        return None

    attempts = db.list_campaign_attempts(campaign_id)
    active_count = sum(1 for item in attempts if item.get("status") == "active")
    completed_count = sum(
        1
        for item in attempts
        if item.get("status") in {"completed", "accepted", "failed"}
    )

    return {
        "campaign": campaign,
        "attempts": attempts,
        "summary": {
            "active": active_count,
            "completed": completed_count,
            "accepted": any(item.get("status") == "accepted" for item in attempts),
        },
    }


def get_blood_campaign_map_points(campaign_id: str) -> dict[str, Any] | None:
    campaign = db.get_blood_campaign(campaign_id)
    if not campaign:
        return None

    donors = db.list_blood_donors(str(campaign["doctor_id"]))
    ngos = db.list_blood_ngos(str(campaign["doctor_id"]))

    attempt_rows = db.list_campaign_attempts(campaign_id)
    attempt_by_donor = {str(row.get("donor_id")): row for row in attempt_rows}

    donor_points = []
    for donor in donors:
        donor_id = str(donor.get("id"))
        attempt = attempt_by_donor.get(donor_id)
        donor_points.append(
            {
                "id": donor_id,
                "name": donor.get("name"),
                "blood_type": donor.get("blood_type"),
                "location": donor.get("location"),
                "latitude": donor.get("latitude"),
                "longitude": donor.get("longitude"),
                "status": attempt.get("status") if attempt else "not-contacted",
            }
        )

    ngo_points = [
        {
            "id": str(ngo.get("id")),
            "name": ngo.get("ngo_name"),
            "location": ngo.get("location"),
            "latitude": ngo.get("latitude"),
            "longitude": ngo.get("longitude"),
        }
        for ngo in ngos
    ]

    patient_point = {
        "location": campaign.get("patient_location"),
        "latitude": campaign.get("patient_latitude"),
        "longitude": campaign.get("patient_longitude"),
    }

    return {
        "campaign_id": campaign_id,
        "patient": patient_point,
        "donors": donor_points,
        "ngos": ngo_points,
    }
